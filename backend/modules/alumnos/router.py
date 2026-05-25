from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import List, Optional
from uuid import UUID
import io
import logging
import openpyxl
from fastapi.responses import StreamingResponse
from fastapi import UploadFile, File
from core.database import get_db
from modules.auth.dependencies import get_current_tenant, TenantContext
from modules.alumnos.schemas import AlumnoCreate, AlumnoUpdate, AlumnoResponse
from modules.alumnos import crud
from modules.alumnos.sync import run_external_sync

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/alumnos", tags=["alumnos"])


@router.get("/template")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos"
    ws.append(["RUT", "Nombre Completo", "Curso"])
    ws.append(["12345678-9", "Juan Pérez González", "1A"])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="planilla_alumnos.xlsx"'
    }
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@router.post("/bulk-upload")
def bulk_upload_alumnos(
    file: UploadFile = File(...),
    target_colegio_id: Optional[str] = None,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para realizar cargas masivas")
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un .xlsx")
    
    # Validar si se sobreescribe el colegio_id (solo para admins)
    colegio_id = tenant.colegio_id
    if target_colegio_id:
        if tenant.rol != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="Solo administradores pueden seleccionar un colegio de destino"
            )
        colegio_id = target_colegio_id

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Omitir cabecera
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            raise HTTPException(status_code=400, detail="El archivo está vacío o no tiene datos.")
            
        data = []
        for row in rows[1:]:
            # Asume columnas: RUT, Nombre Completo, Curso
            if not row or not row[0]:
                continue
            data.append({
                "rut": str(row[0]).strip(),
                "nombre_completo": str(row[1]).strip() if row[1] else "",
                "curso": str(row[2]).strip() if row[2] else ""
            })
            
        stats = crud.bulk_upsert_alumnos(db, data, colegio_id)
        return {"detail": "Importación completada", "stats": stats}
        
    except HTTPException:
        raise
    except Exception:
        logger.exception("Error procesando bulk-upload de alumnos")
        raise HTTPException(status_code=500, detail="Error procesando el archivo")


@router.post("/sync-external")
def sync_external_alumnos(
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant),
):
    if tenant.rol not in ["admin", "coordinador"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permiso para sincronizar con bases de datos externas",
        )
    try:
        results = run_external_sync(db, tenant.colegio_id, tenant.rol)
        return {"detail": "Sincronización externa finalizada", "results": results}
    except Exception as e:
        logger.exception("Error en sincronización externa")
        raise HTTPException(status_code=500, detail=f"Error en sincronización: {str(e)}")


@router.get("", response_model=List[AlumnoResponse])
def list_alumnos(
    skip: int = 0,
    limit: int = 5000,
    taller_id: Optional[UUID] = None,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    return crud.get_alumnos(db, tenant.colegio_id, skip=skip, limit=limit, usuario_id=tenant.usuario_id, rol=tenant.rol, taller_id=taller_id)


@router.get("/stats")
def get_alumnos_stats(
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    return crud.get_alumnos_stats(db, tenant.colegio_id, usuario_id=tenant.usuario_id, rol=tenant.rol)


@router.get("/{alumno_id}", response_model=AlumnoResponse)
def get_alumno(alumno_id: UUID, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    db_alumno = crud.get_alumno_by_id(db, alumno_id, tenant.colegio_id)
    if not db_alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    return db_alumno


@router.post("", response_model=AlumnoResponse, status_code=status.HTTP_201_CREATED)
def create_alumno(alumno: AlumnoCreate, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para crear alumnos")
    existing = crud.get_alumno_by_rut(db, alumno.rut, tenant.colegio_id)
    if existing:
        raise HTTPException(status_code=400, detail="El RUT ya está registrado")
    return crud.create_alumno(db, alumno, tenant.colegio_id)


@router.patch("/{alumno_id}", response_model=AlumnoResponse)
def update_alumno(alumno_id: UUID, alumno: AlumnoUpdate, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para editar alumnos")
    db_alumno = crud.update_alumno(db, alumno_id, alumno, tenant.colegio_id)
    if not db_alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    return db_alumno


@router.delete("/{alumno_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_alumno(alumno_id: UUID, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para eliminar alumnos")
    crud.delete_alumno(db, alumno_id, tenant.colegio_id)

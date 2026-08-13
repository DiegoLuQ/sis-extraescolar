from sqlalchemy.orm import Session
from sqlalchemy import func, extract, distinct
from modules.talleres.models import Taller
from modules.colegios.models import Colegio
from modules.sesiones.models import Sesion
from modules.asistencias.models import Asistencia, EstadoAsistenciaEnum
from modules.inscripciones.models import Inscripcion, EstadoInscripcionEnum
from typing import List, Dict, Optional
import datetime

# Estados que cuentan como "asistió" en el numerador. Solo 'ausente' baja el %.
ESTADOS_PRESENCIA = {
    EstadoAsistenciaEnum.presente,
    EstadoAsistenciaEnum.atraso,
    EstadoAsistenciaEnum.justificado,
}
META_POR_DEFECTO = 85

def get_monthly_attendance_report(db: Session, colegio_id: str, mes: int, anio: int):
    # 1. Obtener todos los talleres del colegio y periodo
    talleres = db.query(Taller).filter(
        Taller.colegio_id == colegio_id,
        Taller.periodo == anio,
        Taller.is_active == True
    ).all()
    
    if not talleres:
        return {
            "active_days": [],
            "workshops": [],
            "daily_stats": {},
            "total_enrollment": 0
        }
        
    taller_ids = [t.id for t in talleres]
    
    # 2. Obtener sesiones del mes (excluyendo fines de semana)
    sesiones = db.query(Sesion).filter(
        Sesion.taller_id.in_(taller_ids),
        extract('month', Sesion.fecha_sesion) == mes,
        extract('year', Sesion.fecha_sesion) == anio,
        func.dayofweek(Sesion.fecha_sesion).notin_([1, 7])
    ).all()
    
    # 3. Obtener matrícula por taller
    enrollment_query = db.query(Inscripcion.taller_id, func.count(Inscripcion.id).label("count")).filter(
        Inscripcion.taller_id.in_(taller_ids),
        Inscripcion.estado == EstadoInscripcionEnum.inscrito
    ).group_by(Inscripcion.taller_id).all()
    enrollment_map = {str(r.taller_id): r.count for r in enrollment_query}
    total_enrollment = sum(enrollment_map.values())
    
    # 4. Obtener datos de asistencia por sesión
    asistencias_raw = db.query(
        Asistencia.sesion_id,
        func.count(Asistencia.id).label("presentes")
    ).filter(
        Asistencia.sesion_id.in_([s.id for s in sesiones]),
        Asistencia.estado_asistencia == "presente"
    ).group_by(Asistencia.sesion_id).all()
    asistencias_map = {str(r.sesion_id): r.presentes for r in asistencias_raw}
    
    # 5. Organizar datos para la matriz
    active_days = sorted(list(set([s.fecha_sesion.strftime("%Y-%m-%d") for s in sesiones])))
    workshops_report = []
    
    for t in talleres:
        matriculados_taller = enrollment_map.get(str(t.id), 0)
        taller_asistencias = {}
        total_taller = 0
        for day in active_days:
            sesiones_hoy = [s for s in sesiones if str(s.taller_id) == str(t.id) and s.fecha_sesion.strftime("%Y-%m-%d") == day]
            if sesiones_hoy:
                count = sum(asistencias_map.get(str(s.id), 0) for s in sesiones_hoy)
                taller_asistencias[day] = count
                total_taller += count
            else:
                taller_asistencias[day] = None
        
        workshops_report.append({
            "id": t.id,
            "nombre_taller": t.nombre_taller,
            "matriculados": matriculados_taller,
            "asistencias": taller_asistencias,
            "total_taller": total_taller
        })
        
    # 6. Estadísticas diarias
    daily_stats = {}
    for day in active_days:
        presentes_dia = sum([w["asistencias"].get(day) or 0 for w in workshops_report])
        capacidad_dia = sum([w["matriculados"] for w in workshops_report if w["asistencias"].get(day) is not None])
        porcentaje = (presentes_dia / capacidad_dia * 100) if capacidad_dia > 0 else 0
        daily_stats[day] = {
            "presentes": presentes_dia,
            "matricula_total": capacidad_dia,
            "porcentaje": round(porcentaje, 1)
        }
        
    return {
        "active_days": active_days,
        "workshops": workshops_report,
        "daily_stats": daily_stats,
        "total_enrollment": total_enrollment
    }

def get_weekly_attendance_report(db: Session, colegio_id: str, mes: int, anio: int):
    talleres = db.query(Taller).filter(Taller.colegio_id == colegio_id, Taller.periodo == anio).all()
    taller_ids = [t.id for t in talleres]
    
    # 1. Matrícula por taller
    enrollment_query = db.query(Inscripcion.taller_id, func.count(Inscripcion.id).label("count")).filter(
        Inscripcion.taller_id.in_(taller_ids),
        Inscripcion.estado == EstadoInscripcionEnum.inscrito
    ).group_by(Inscripcion.taller_id).all()
    enrollment_map = {str(r.taller_id): r.count for r in enrollment_query}
    total_enrollment = sum(enrollment_map.values())

    # 2. Sesiones del mes
    sesiones = db.query(Sesion).filter(
        Sesion.colegio_id == colegio_id,
        extract('month', Sesion.fecha_sesion) == mes,
        extract('year', Sesion.fecha_sesion) == anio
    ).all()
    
    # 3. Asistencias presentes por sesión
    asistencias_raw = db.query(
        Asistencia.sesion_id,
        func.count(Asistencia.id).label("presentes")
    ).filter(
        Asistencia.sesion_id.in_([s.id for s in sesiones]),
        Asistencia.estado_asistencia == "presente"
    ).group_by(Asistencia.sesion_id).all()
    asistencias_map = {str(r.sesion_id): r.presentes for r in asistencias_raw}

    # 4. Agrupar por semana
    from itertools import groupby
    def get_week_label(s):
        return s.fecha_sesion.isocalendar()[1]
    
    sesiones_sorted = sorted(sesiones, key=lambda s: s.fecha_sesion)
    weeks_data = []
    
    for i, (week_num, group) in enumerate(groupby(sesiones_sorted, key=get_week_label)):
        group_list = list(group)
        presentes = sum([asistencias_map.get(str(s.id), 0) for s in group_list])
        capacidad = sum([enrollment_map.get(str(s.taller_id), 0) for s in group_list])
        porcentaje = (presentes / capacidad * 100) if capacidad > 0 else 0
        
        weeks_data.append({
            "week_label": f"Semana {i+1}",
            "presentes": presentes,
            "capacidad": capacidad,
            "porcentaje": round(min(porcentaje, 100), 1)
        })

    return {
        "weeks": weeks_data,
        "total_enrollment": total_enrollment,
        "active_days_count": len(set([s.fecha_sesion for s in sesiones]))
    }

def get_annual_attendance_report(db: Session, colegio_id: str, anio: int):
    talleres = db.query(Taller).filter(Taller.colegio_id == colegio_id, Taller.periodo == anio).all()
    taller_ids = [t.id for t in talleres]
    
    # 1. Matrícula por taller
    enrollment_query = db.query(Inscripcion.taller_id, func.count(Inscripcion.id).label("count")).filter(
        Inscripcion.taller_id.in_(taller_ids),
        Inscripcion.estado == EstadoInscripcionEnum.inscrito
    ).group_by(Inscripcion.taller_id).all()
    enrollment_map = {str(r.taller_id): r.count for r in enrollment_query}
    total_enrollment = sum(enrollment_map.values())

    # 2. Todas las sesiones del año
    sesiones = db.query(Sesion).filter(
        Sesion.colegio_id == colegio_id,
        extract('year', Sesion.fecha_sesion) == anio
    ).all()
    
    # 3. Asistencias presentes por sesión
    asistencias_raw = db.query(
        Asistencia.sesion_id,
        func.count(Asistencia.id).label("presentes")
    ).filter(
        Asistencia.sesion_id.in_([s.id for s in sesiones]),
        Asistencia.estado_asistencia == "presente"
    ).group_by(Asistencia.sesion_id).all()
    asistencias_map = {str(r.sesion_id): r.presentes for r in asistencias_raw}

    meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    months_data = []
    
    total_presentes_anio = 0
    total_capacidad_anio = 0
    
    for m_idx in range(1, 13):
        mes_sesiones = [s for s in sesiones if s.fecha_sesion.month == m_idx]
        if not mes_sesiones:
            continue
            
        presentes = sum([asistencias_map.get(str(s.id), 0) for s in mes_sesiones])
        capacidad = sum([enrollment_map.get(str(s.taller_id), 0) for s in mes_sesiones])
        porcentaje = (presentes / capacidad * 100) if capacidad > 0 else 0
        
        total_presentes_anio += presentes
        total_capacidad_anio += capacidad
        
        months_data.append({
            "month_label": meses_nombres[m_idx - 1],
            "presentes": presentes,
            "capacidad": capacidad,
            "porcentaje": round(min(porcentaje, 100), 1)
        })

    return {
        "months": months_data,
        "total_enrollment": total_enrollment,
        "total_presentes_anio": total_presentes_anio,
        "promedio_anual": round((total_presentes_anio / total_capacidad_anio * 100), 1) if total_capacidad_anio > 0 else 0
    }

def get_metas_report(db: Session, colegio_id: str, mes: int, anio: int):
    """Reporte de cumplimiento de metas por taller (semanal y mensual).

    El % se calcula SOBRE LA LISTA REGISTRADA: denominador = todos los registros de
    asistencia de las sesiones del rango; numerador = registros en estado
    presente/atraso/justificado (solo 'ausente' baja el %). Se excluyen los registros
    de alumnos cuya fecha de retiro del taller es anterior a la fecha de la sesión,
    para conservar el histórico exacto.
    """
    colegio = db.query(Colegio).filter(Colegio.id == colegio_id).first()
    colegio_meta = (colegio.meta_asistencia if colegio and colegio.meta_asistencia else META_POR_DEFECTO)

    talleres = db.query(Taller).filter(
        Taller.colegio_id == colegio_id,
        Taller.periodo == anio,
        Taller.is_active == True
    ).all()
    taller_ids = [t.id for t in talleres]

    sesiones = []
    if taller_ids:
        sesiones = db.query(Sesion).filter(
            Sesion.taller_id.in_(taller_ids),
            extract('month', Sesion.fecha_sesion) == mes,
            extract('year', Sesion.fecha_sesion) == anio
        ).all()

    # Mapa sesion_id -> (taller_id, fecha)
    sesion_info = {str(s.id): (str(s.taller_id), s.fecha_sesion) for s in sesiones}
    sesion_ids = list(sesion_info.keys())

    # Asistencias de esas sesiones
    asistencias = []
    if sesion_ids:
        asistencias = db.query(
            Asistencia.sesion_id, Asistencia.alumno_id, Asistencia.estado_asistencia
        ).filter(Asistencia.sesion_id.in_(sesion_ids)).all()

    # Mapa de fechas de retiro: (alumno_id, taller_id) -> fecha_retiro (solo retirados)
    retiro_map = {}
    if taller_ids:
        retiros = db.query(
            Inscripcion.alumno_id, Inscripcion.taller_id, Inscripcion.fecha_retiro
        ).filter(
            Inscripcion.taller_id.in_(taller_ids),
            Inscripcion.estado == EstadoInscripcionEnum.retirado,
            Inscripcion.fecha_retiro.isnot(None)
        ).all()
        retiro_map = {(str(r.alumno_id), str(r.taller_id)): r.fecha_retiro for r in retiros}

    def _build(sesiones_subset):
        subset_ids = {str(s.id) for s in sesiones_subset}
        # acumuladores por taller
        acc = {tid: {"asistencias": 0, "registros": 0} for tid in [str(t.id) for t in talleres]}
        for sesion_id, alumno_id, estado in asistencias:
            sid = str(sesion_id)
            if sid not in subset_ids:
                continue
            taller_id, fecha = sesion_info[sid]
            # Excluir registros posteriores al retiro del alumno en ese taller
            retiro = retiro_map.get((str(alumno_id), taller_id))
            if retiro and retiro < fecha:
                continue
            if taller_id not in acc:
                continue
            acc[taller_id]["registros"] += 1
            if estado in ESTADOS_PRESENCIA:
                acc[taller_id]["asistencias"] += 1

        items = []
        tot_asist = 0
        tot_reg = 0
        for t in talleres:
            tid = str(t.id)
            a = acc[tid]["asistencias"]
            r = acc[tid]["registros"]
            tot_asist += a
            tot_reg += r
            meta = t.meta_asistencia if t.meta_asistencia else colegio_meta
            pct = round((a / r) * 100, 1) if r > 0 else 0.0
            items.append({
                "taller_id": t.id,
                "nombre_taller": t.nombre_taller,
                "asistencias": a,
                "registros": r,
                "porcentaje": pct,
                "meta": meta,
                "cumple": r > 0 and pct >= meta,
            })
        # Ordenar por asistencia (nº) descendente; desempate por porcentaje
        items.sort(key=lambda x: (x["asistencias"], x["porcentaje"]), reverse=True)
        total_pct = round((tot_asist / tot_reg) * 100, 1) if tot_reg > 0 else 0.0
        total = {
            "asistencias": tot_asist,
            "registros": tot_reg,
            "porcentaje": total_pct,
            "meta": colegio_meta,
            "cumple": tot_reg > 0 and total_pct >= colegio_meta,
        }
        return items, total

    mensual, total_mensual = _build(sesiones)

    # Agrupar las sesiones del mes por semana ISO (para navegar semana a semana)
    semanas_map = {}
    for s in sesiones:
        wk = s.fecha_sesion.isocalendar()[1]
        semanas_map.setdefault(wk, []).append(s)

    semanas = []
    for wk in sorted(semanas_map.keys()):
        grupo = semanas_map[wk]
        items_s, total_s = _build(grupo)
        fechas = [s.fecha_sesion for s in grupo]
        semanas.append({
            "semana_inicio": min(fechas).isoformat(),
            "semana_fin": max(fechas).isoformat(),
            "items": items_s,
            "total": total_s,
        })

    return {
        "mensual": mensual,
        "total_mensual": total_mensual,
        "semanas": semanas,
    }


def get_weekly_summary_report(db: Session, colegio_id: str, semanas_atras: int = 0):
    hoy = datetime.date.today()
    
    # Calcular el lunes de la semana actual (hoy - su día de la semana)
    # hoy.weekday() es 0 para lunes, 6 para domingo
    inicio_semana_actual = hoy - datetime.timedelta(days=hoy.weekday())
    
    # Aplicar el offset de semanas
    inicio_semana_seleccionada = inicio_semana_actual - datetime.timedelta(weeks=semanas_atras)
    fin_semana_seleccionada = inicio_semana_seleccionada + datetime.timedelta(days=4)  # Lunes a Viernes
    
    # Semana anterior relativa a la seleccionada
    inicio_semana_anterior = inicio_semana_seleccionada - datetime.timedelta(weeks=1)
    fin_semana_anterior = inicio_semana_anterior + datetime.timedelta(days=4)  # Lunes a Viernes
    
    # Rango extendido para incluir el día anterior (el domingo anterior)
    rango_sel_inicio = inicio_semana_seleccionada - datetime.timedelta(days=1)
    rango_ant_inicio = inicio_semana_anterior - datetime.timedelta(days=1)
    
    # Obtener talleres activos del colegio
    talleres = db.query(Taller).filter(Taller.colegio_id == colegio_id, Taller.is_active == True).all()
    taller_ids = [t.id for t in talleres]

    if not taller_ids:
        return {
            "semana_actual": [],
            "comparativa_dia_anterior": None,
            "total_hoy": 0,
            "total_semana_actual": 0,
            "total_semana_anterior": 0,
            "comparativa_totales_porcentaje": None
        }

    # Sesiones en las semanas respectivas para talleres activos
    sesiones_seleccionadas = db.query(Sesion).filter(
        Sesion.colegio_id == colegio_id,
        Sesion.taller_id.in_(taller_ids),
        Sesion.fecha_sesion >= rango_sel_inicio,
        Sesion.fecha_sesion <= fin_semana_seleccionada
    ).all()

    sesiones_anteriores = db.query(Sesion).filter(
        Sesion.colegio_id == colegio_id,
        Sesion.taller_id.in_(taller_ids),
        Sesion.fecha_sesion >= rango_ant_inicio,
        Sesion.fecha_sesion <= fin_semana_anterior
    ).all()

    sesiones_hoy = db.query(Sesion).filter(
        Sesion.colegio_id == colegio_id,
        Sesion.taller_id.in_(taller_ids),
        Sesion.fecha_sesion == hoy
    ).all()
    enrollment_map = {}
    if taller_ids:
        enrollment_query = db.query(Inscripcion.taller_id, func.count(Inscripcion.id).label("count")).filter(
            Inscripcion.taller_id.in_(taller_ids),
            Inscripcion.estado == EstadoInscripcionEnum.inscrito
        ).group_by(Inscripcion.taller_id).all()
        enrollment_map = {str(r.taller_id): r.count for r in enrollment_query}
    
    def get_attendance_for_sessions(sesion_ids):
        if not sesion_ids:
            return {}
        res = db.query(
            Asistencia.sesion_id,
            func.count(Asistencia.id).label("presentes")
        ).filter(
            Asistencia.sesion_id.in_(sesion_ids),
            Asistencia.estado_asistencia == "presente"
        ).group_by(Asistencia.sesion_id).all()
        return {str(row.sesion_id): row.presentes for row in res}
    
    asistencias_sel_map = get_attendance_for_sessions([s.id for s in sesiones_seleccionadas])
    asistencias_ant_map = get_attendance_for_sessions([s.id for s in sesiones_anteriores])
    asistencias_hoy_map = get_attendance_for_sessions([s.id for s in sesiones_hoy])
    
    total_hoy = sum(asistencias_hoy_map.values())
    
    dias_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    
    def build_weekly_list(inicio_lunes, sesiones_lista, asistencias_map):
        days_list = []
        date_to_sessions = {}
        sesion_obj_map = {s.id: s for s in sesiones_lista}
        
        for s in sesiones_lista:
            dt_str = s.fecha_sesion.strftime("%Y-%m-%d")
            if dt_str not in date_to_sessions:
                date_to_sessions[dt_str] = []
            date_to_sessions[dt_str].append(s.id)
            
        # Porcentaje de asistencia del último día CON actividad (sesiones), para comparar
        # contra él aunque el día calendario inmediatamente anterior no haya tenido talleres.
        porcentaje_ultimo_activo = None

        for i in range(5):  # Lunes a Viernes
            day_date = inicio_lunes + datetime.timedelta(days=i)
            day_date_str = day_date.strftime("%Y-%m-%d")
            session_ids = date_to_sessions.get(day_date_str, [])
            presentes_dia = sum(asistencias_map.get(sid, 0) for sid in session_ids)

            # Calcular matrícula de los talleres con sesiones hoy
            unique_taller_ids = set(sesion_obj_map[sid].taller_id for sid in session_ids if sid in sesion_obj_map)
            matricula_total_dia = sum(enrollment_map.get(str(tid), 0) for tid in unique_taller_ids)

            porcentaje_asistencia = round((presentes_dia / matricula_total_dia) * 100, 1) if matricula_total_dia > 0 else 0.0

            # Variación de asistencia en PUNTOS PORCENTUALES (pp) respecto al último día con
            # actividad, no como porcentaje relativo. Ej: 58.3% → 79.2% = +20.9 pp.
            # Si el día anterior del calendario no tuvo talleres, se compara contra el último
            # día que sí tuvo. El primer día con actividad de la semana no tiene referencia.
            diferencia_anterior = None
            if matricula_total_dia > 0:
                if porcentaje_ultimo_activo is not None:
                    diferencia_anterior = round(porcentaje_asistencia - porcentaje_ultimo_activo, 1)
                porcentaje_ultimo_activo = porcentaje_asistencia

            days_list.append({
                "fecha": day_date_str,
                "dia": dias_nombres[i],
                "presentes": presentes_dia,
                "matricula_total": matricula_total_dia,
                "porcentaje_asistencia": porcentaje_asistencia,
                "diferencia_anterior": diferencia_anterior
            })
        return days_list
        
    semana_actual_list = build_weekly_list(inicio_semana_seleccionada, sesiones_seleccionadas, asistencias_sel_map)
    semana_anterior_list = build_weekly_list(inicio_semana_anterior, sesiones_anteriores, asistencias_ant_map)
    
    total_semana_actual = sum(d["presentes"] for d in semana_actual_list)
    total_semana_anterior = sum(d["presentes"] for d in semana_anterior_list)
    
    comparativa_totales_porcentaje = None
    if total_semana_anterior > 0:
        comparativa_totales_porcentaje = round(((total_semana_actual - total_semana_anterior) / total_semana_anterior) * 100, 1)
        
    return {
        "total_hoy": total_hoy,
        "semana_actual": semana_actual_list,
        "semana_anterior": semana_anterior_list,
        "total_semana_actual": total_semana_actual,
        "total_semana_anterior": total_semana_anterior,
        "comparativa_totales_porcentaje": comparativa_totales_porcentaje
    }


def generate_attendance_excel(
    db: Session,
    colegio_id: str,
    mes: int,
    anio: int,
    fecha_inicio: str = None,
    fecha_fin: str = None,
    taller_id: str = None,
    dias_semana_str: str = None,
    export_modo: str = "mes"
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    colegio = db.query(Colegio).filter(Colegio.id == colegio_id).first()
    nombre_colegio = colegio.nombre_colegio if colegio else "Colegio"
    colegio_lower = nombre_colegio.lower()

    # Definir paleta temática
    if "macaya" in colegio_lower:
        # Tema Macaya: Verde oscuro / esmeralda
        primary_color = "14532D"      # Dark green fill
        accent_color = "DCFCE7"       # Soft green tint fill
        header_font_color = "FFFFFF"  # White text
        border_color = "A7F3D0"       # Light green border
    else:
        # Tema DP (Diego Portales / Defecto): Azul oscuro
        primary_color = "1E3A8A"      # Dark blue fill
        accent_color = "DBEAFE"       # Soft blue tint fill
        header_font_color = "FFFFFF"  # White text
        border_color = "BFDBFE"       # Light blue border

    # 1. Obtener talleres
    talleres_q = db.query(Taller).filter(
        Taller.colegio_id == colegio_id,
        Taller.periodo == anio,
        Taller.is_active == True
    )
    if taller_id:
        talleres_q = talleres_q.filter(Taller.id == taller_id)
    talleres = talleres_q.all()

    if not talleres:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="No hay talleres registrados para este período")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    taller_ids = [t.id for t in talleres]

    # 2. Obtener sesiones según el modo de exportación
    sesiones_q = db.query(Sesion).filter(
        Sesion.taller_id.in_(taller_ids),
        func.dayofweek(Sesion.fecha_sesion).notin_([1, 7])
    )

    if export_modo == "anio":
        sesiones_q = sesiones_q.filter(extract('year', Sesion.fecha_sesion) == anio)
    elif export_modo == "rango" and fecha_inicio and fecha_fin:
        sesiones_q = sesiones_q.filter(
            func.date(Sesion.fecha_sesion) >= fecha_inicio,
            func.date(Sesion.fecha_sesion) <= fecha_fin
        )
    else:
        sesiones_q = sesiones_q.filter(
            extract('month', Sesion.fecha_sesion) == mes,
            extract('year', Sesion.fecha_sesion) == anio
        )

    sesiones = sesiones_q.all()

    # 3. Matrícula por taller
    enrollment_query = db.query(Inscripcion.taller_id, func.count(Inscripcion.id).label("count")).filter(
        Inscripcion.taller_id.in_(taller_ids),
        Inscripcion.estado == EstadoInscripcionEnum.inscrito
    ).group_by(Inscripcion.taller_id).all()
    enrollment_map = {str(r.taller_id): r.count for r in enrollment_query}

    # 4. Asistencias presentes por sesión
    sesion_ids = [s.id for s in sesiones]
    asistencias_map = {}
    if sesion_ids:
        asistencias_raw = db.query(
            Asistencia.sesion_id,
            func.count(Asistencia.id).label("presentes")
        ).filter(
            Asistencia.sesion_id.in_(sesion_ids),
            Asistencia.estado_asistencia == "presente"
        ).group_by(Asistencia.sesion_id).all()
        asistencias_map = {str(r.sesion_id): r.presentes for r in asistencias_raw}

    # 5. Organizar días activos
    active_days = sorted(list(set([s.fecha_sesion.strftime("%Y-%m-%d") for s in sesiones])))

    dias_nombres = ["Domingo", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

    def get_nombre_dia(fecha_str):
        d = datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
        return dias_nombres[d.weekday() if d.weekday() != 6 else 0] if d.weekday() == 6 else dias_nombres[d.weekday() + 1]

    dias_seleccionados = dias_semana_str.split(",") if dias_semana_str else ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    # Filtrar días según rango y días de semana si aplica
    active_days_filtrados = []
    for day in active_days:
        if export_modo == "mes":
            if fecha_inicio and day < fecha_inicio:
                continue
            if fecha_fin and day > fecha_fin:
                continue
        nombre_dia = get_nombre_dia(day)
        if dias_seleccionados and nombre_dia not in dias_seleccionados:
            continue
        active_days_filtrados.append(day)

    # Construir mapa de asistencias por taller y día
    workshops_filtrados = []
    for t in talleres:
        matriculados_taller = enrollment_map.get(str(t.id), 0)
        taller_asistencias = {}
        total_taller = 0
        for day in active_days_filtrados:
            sesiones_hoy = [s for s in sesiones if str(s.taller_id) == str(t.id) and s.fecha_sesion.strftime("%Y-%m-%d") == day]
            if sesiones_hoy:
                count = sum(asistencias_map.get(str(s.id), 0) for s in sesiones_hoy)
                taller_asistencias[day] = count
                total_taller += count
            else:
                taller_asistencias[day] = None

        if active_days_filtrados:
            tiene_actividad = any(taller_asistencias.get(day) is not None for day in active_days_filtrados)
            if not tiene_actividad and export_modo == "mes":
                continue

        workshops_filtrados.append({
            "id": t.id,
            "nombre_taller": t.nombre_taller,
            "matriculados": matriculados_taller,
            "asistencias": taller_asistencias,
            "total_taller": total_taller
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle Asistencia"
    ws.views.sheetView[0].showGridLines = True

    # Estilos openpyxl
    title_font = Font(name="Calibri", size=16, bold=True, color=primary_color)
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="555555")
    header_font = Font(name="Calibri", size=11, bold=True, color=header_font_color)
    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    
    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    summary_fill = PatternFill(start_color=accent_color, end_color=accent_color, fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Fila 1 y 2: Título
    meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    mes_str = meses_nombres[mes - 1] if 1 <= mes <= 12 else str(mes)

    if export_modo == "anio":
        periodo_texto = f"Año Completo {anio}"
    elif export_modo == "rango" and fecha_inicio and fecha_fin:
        periodo_texto = f"Rango: {fecha_inicio} a {fecha_fin} ({anio})"
    else:
        periodo_texto = f"Mes de {mes_str} {anio}"

    ws.cell(row=1, column=1, value=f"REPORTE DETALLE DE ASISTENCIA POR TALLER - {nombre_colegio.upper()}").font = title_font
    ws.cell(row=2, column=1, value=f"Período: {periodo_texto} | Exportado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}").font = subtitle_font

    # Fila 4: Encabezados
    start_row = 4
    headers = ["Taller", "Inscritos"]
    meses_abrev = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
    for day in active_days_filtrados:
        parts = day.split("-")
        mes_num = int(parts[1])
        dia_num = parts[2]
        dia_nom = get_nombre_dia(day)[:3].upper()
        
        if export_modo in ["anio", "rango"]:
            mes_abrev = meses_abrev[mes_num - 1]
            headers.append(f"{dia_nom} {dia_num}/{mes_abrev}")
        else:
            headers.append(f"{dia_nom} {dia_num}")
    headers.extend(["Total", "Promedio"])

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = grid_border

    ws.row_dimensions[start_row].height = 26

    # Filas de datos
    current_row = start_row + 1
    for r_idx, w in enumerate(workshops_filtrados):
        ws.row_dimensions[current_row].height = 20
        row_fill = zebra_fill if r_idx % 2 == 1 else None

        # Taller
        c_taller = ws.cell(row=current_row, column=1, value=w["nombre_taller"])
        c_taller.font = bold_font
        c_taller.alignment = align_left
        c_taller.border = grid_border
        if row_fill: c_taller.fill = row_fill

        # Inscritos
        c_ins = ws.cell(row=current_row, column=2, value=w["matriculados"])
        c_ins.font = data_font
        c_ins.alignment = align_center
        c_ins.border = grid_border
        if row_fill: c_ins.fill = row_fill

        # Días
        total_taller = 0
        sesiones_contadas = 0
        for d_idx, day in enumerate(active_days_filtrados, 3):
            val = w["asistencias"].get(day)
            c_val = ws.cell(row=current_row, column=d_idx)
            if val is not None:
                c_val.value = val
                total_taller += val
                sesiones_contadas += 1
            else:
                c_val.value = "-"
            c_val.font = data_font
            c_val.alignment = align_center
            c_val.border = grid_border
            if row_fill: c_val.fill = row_fill

        prom = round(total_taller / sesiones_contadas, 1) if sesiones_contadas > 0 else 0.0

        # Total
        col_total_idx = len(headers) - 1
        c_tot = ws.cell(row=current_row, column=col_total_idx, value=total_taller)
        c_tot.font = bold_font
        c_tot.fill = summary_fill
        c_tot.alignment = align_center
        c_tot.border = grid_border

        # Promedio
        c_prom = ws.cell(row=current_row, column=len(headers), value=prom)
        c_prom.font = bold_font
        c_prom.fill = summary_fill
        c_prom.alignment = align_center
        c_prom.border = grid_border

        current_row += 1

    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < start_row:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 11) if col_letter != "A" else max(max_len + 4, 28)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"detalle_asistencia_{nombre_colegio}_{mes_str}_{anio}.xlsx".replace(" ", "_")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def get_ranking_alumnos(
    db: Session,
    colegio_id: str,
    mes: Optional[int] = None,
    anio: Optional[int] = None,
    taller_id: Optional[str] = None
) -> List[Dict]:
    """
    Retorna la lista de alumnos ordenados por su porcentaje global de asistencia
    en sus talleres correspondientes al colegio especificado.
    """
    from modules.alumnos.models import Alumno
    from modules.inscripciones.models import Inscripcion
    from modules.talleres.models import Taller
    from modules.sesiones.models import Sesion
    from modules.asistencias.models import Asistencia

    # Query base de alumnos inscritos en talleres del colegio
    query_ins = db.query(Inscripcion, Alumno, Taller).join(
        Alumno, Inscripcion.alumno_id == Alumno.id
    ).join(
        Taller, Inscripcion.taller_id == Taller.id
    ).filter(
        Taller.colegio_id == str(colegio_id),
        Inscripcion.estado == "inscrito"
    )

    if taller_id:
        query_ins = query_ins.filter(Taller.id == str(taller_id))

    inscripciones = query_ins.all()
    if not inscripciones:
        return []

    # Mapear datos por alumno
    alumnos_dict = {}
    for ins, alu, tal in inscripciones:
        if alu.id not in alumnos_dict:
            alumnos_dict[alu.id] = {
                "alumno_id": alu.id,
                "nombre_completo": alu.nombre_completo,
                "rut": alu.rut,
                "curso": alu.curso,
                "talleres": [tal.nombre_taller],
                "taller_ids": [tal.id],
                "presentes": 0,
                "ausentes": 0,
                "justificados": 0,
                "atrasos": 0,
                "total_sesiones": 0
            }
        else:
            if tal.nombre_taller not in alumnos_dict[alu.id]["talleres"]:
                alumnos_dict[alu.id]["talleres"].append(tal.nombre_taller)
                alumnos_dict[alu.id]["taller_ids"].append(tal.id)

    alumno_ids = list(alumnos_dict.keys())

    # Obtener todas las asistencias de sesiones realizadas
    query_asist = db.query(Asistencia, Sesion).join(
        Sesion, Asistencia.sesion_id == Sesion.id
    ).join(
        Taller, Sesion.taller_id == Taller.id
    ).filter(
        Asistencia.alumno_id.in_(alumno_ids),
        Taller.colegio_id == str(colegio_id)
    )

    if taller_id:
        query_asist = query_asist.filter(Taller.id == str(taller_id))
    if anio:
        query_asist = query_asist.filter(extract('year', Sesion.fecha_sesion) == anio)
    if mes:
        query_asist = query_asist.filter(extract('month', Sesion.fecha_sesion) == mes)

    asistencias_list = query_asist.all()

    for asist, ses in asistencias_list:
        alu_id = asist.alumno_id
        if alu_id in alumnos_dict:
            st = (asist.estado_asistencia or "").lower()
            alumnos_dict[alu_id]["total_sesiones"] += 1
            if st == "presente":
                alumnos_dict[alu_id]["presentes"] += 1
            elif st == "atraso":
                alumnos_dict[alu_id]["presentes"] += 1
                alumnos_dict[alu_id]["atrasos"] += 1
            elif st == "ausente":
                alumnos_dict[alu_id]["ausentes"] += 1
            elif st == "justificado":
                alumnos_dict[alu_id]["justificados"] += 1

    resultado = []
    for alu_id, data in alumnos_dict.items():
        tot = data["total_sesiones"]
        pres = data["presentes"]
        pct = round((pres / tot * 100), 1) if tot > 0 else 0.0
        
        resultado.append({
            "alumno_id": data["alumno_id"],
            "nombre_completo": data["nombre_completo"],
            "rut": data["rut"],
            "curso": data["curso"],
            "talleres": data["talleres"],
            "total_sesiones": tot,
            "presentes": pres,
            "ausentes": data["ausentes"],
            "justificados": data["justificados"],
            "atrasos": data["atrasos"],
            "porcentaje_asistencia": pct
        })

    # Ordenar por porcentaje de asistencia descendente y luego por nombre
    resultado.sort(key=lambda x: (-x["porcentaje_asistencia"], x["nombre_completo"]))
    return resultado


def get_alumno_detalle_asistencia(
    db: Session,
    alumno_id: str,
    colegio_id: str,
    mes: Optional[int] = None,
    anio: Optional[int] = None
) -> Dict:
    """
    Retorna la ficha de asistencia detallada de un alumno, incluyendo
    sus talleres y el listado cronológico de sesiones asistidas/faltadas.
    """
    from modules.alumnos.models import Alumno
    from modules.inscripciones.models import Inscripcion
    from modules.talleres.models import Taller
    from modules.sesiones.models import Sesion
    from modules.asistencias.models import Asistencia

    alumno = db.query(Alumno).filter(Alumno.id == str(alumno_id)).first()
    if not alumno:
        return {"status": "not_found"}

    # Obtener inscripciones
    from modules.usuarios.models import Usuario
    inscripciones = db.query(Inscripcion, Taller, Usuario).join(
        Taller, Inscripcion.taller_id == Taller.id
    ).outerjoin(
        Usuario, Taller.profesor_id == Usuario.id
    ).filter(
        Inscripcion.alumno_id == str(alumno_id),
        Taller.colegio_id == str(colegio_id)
    ).all()

    talleres_list = []
    for ins, tal, usu in inscripciones:
        nombre_prof = (usu.nombre_2 or usu.nombre) if usu else ""
        talleres_list.append({
            "taller_id": tal.id,
            "nombre_taller": tal.nombre_taller,
            "profesor": nombre_prof,
            "estado_inscripcion": ins.estado
        })

    # Obtener historial de asistencias
    query = db.query(Asistencia, Sesion, Taller).join(
        Sesion, Asistencia.sesion_id == Sesion.id
    ).join(
        Taller, Sesion.taller_id == Taller.id
    ).filter(
        Asistencia.alumno_id == str(alumno_id),
        Taller.colegio_id == str(colegio_id)
    )

    if anio:
        query = query.filter(extract('year', Sesion.fecha_sesion) == anio)
    if mes:
        query = query.filter(extract('month', Sesion.fecha_sesion) == mes)

    query = query.order_by(Sesion.fecha_sesion.desc())
    registros = query.all()

    historial = []
    presentes = 0
    ausentes = 0
    justificados = 0
    atrasos = 0

    for asist, ses, tal in registros:
        st = (asist.estado_asistencia or "").lower()
        if st == "presente":
            presentes += 1
        elif st == "atraso":
            presentes += 1
            atrasos += 1
        elif st == "ausente":
            ausentes += 1
        elif st == "justificado":
            justificados += 1

        historial.append({
            "asistencia_id": asist.id,
            "fecha": ses.fecha_sesion.strftime("%Y-%m-%d"),
            "taller_nombre": tal.nombre_taller,
            "estado": asist.estado_asistencia,
            "observaciones": asist.observaciones,
            "bloqueada": ses.bloqueada
        })

    total_sesiones = len(registros)
    porcentaje = round((presentes / total_sesiones * 100), 1) if total_sesiones > 0 else 0.0

    return {
        "alumno": {
            "id": alumno.id,
            "nombre_completo": alumno.nombre_completo,
            "rut": alumno.rut,
            "curso": alumno.curso,
            "telefono": alumno.telefono
        },
        "resumen": {
            "total_sesiones": total_sesiones,
            "presentes": presentes,
            "ausentes": ausentes,
            "justificados": justificados,
            "atrasos": atrasos,
            "porcentaje_asistencia": porcentaje
        },
        "talleres": talleres_list,
        "historial": historial
    }




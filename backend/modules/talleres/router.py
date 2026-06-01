from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from uuid import UUID
import io
import logging
import openpyxl
from fastapi.responses import StreamingResponse
from core.database import get_db
from modules.auth.dependencies import get_current_tenant, TenantContext
from modules.talleres.schemas import TallerCreate, TallerUpdate, TallerResponse, TallerClonacion
from modules.talleres import crud

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/talleres", tags=["talleres"])


@router.get("/template")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Talleres"
    
    # Cabeceras completas
    headers_list = ["Nombre del Taller", "ID del Monitor", "Cupos Máximos", "Día", "Hora Inicio", "Hora Fin", "Cursos Asignados", "Año Académico", "ID del Colegio"]
    ws.append(headers_list)
    
    # Ejemplo 1: Taller con múltiples horarios usando ID de Monitor (UUID) y dejando ID del Colegio en blanco para usar el seleccionado por defecto
    ws.append(["Básquetbol", "ed3d24ac-e49f-4593-ba5c-e2c7010f10d7", 30, "Lunes", "15:00", "16:30", "5ºA, 5ºB, 6ºA", 2026, ""])
    ws.append(["Básquetbol", "ed3d24ac-e49f-4593-ba5c-e2c7010f10d7", 30, "Miércoles", "15:30", "17:00", "5ºA, 5ºB, 6ºA", 2026, ""])
    ws.append(["Básquetbol", "ed3d24ac-e49f-4593-ba5c-e2c7010f10d7", 30, "Viernes", "16:00", "17:30", "5ºA, 5ºB, 6ºA", 2026, ""])
    
    # Ejemplo 2: Taller con un solo horario usando Nombre de Usuario (Login) del monitor
    ws.append(["Ajedrez Básica", "tenis_de_mesa_1", 20, "Martes", "14:00", "15:30", "1ºA, 1ºB, 2ºA", 2026, ""])
    
    # Estilo básico para que se vea premium
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="planilla_talleres.xlsx"'
    }
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.post("/bulk-upload")
def bulk_upload_talleres(
    file: UploadFile = File(...),
    target_colegio_id: Optional[str] = None,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para realizar cargas masivas")
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un .xlsx")
    
    # Validar si se sobreescribe el colegio_id (solo para admins)
    colegio_id = tenant.colegio_id
    if target_colegio_id:
        if tenant.rol != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="Solo administradores pueden seleccionar un colegio de destino"
            )
        colegio_id = target_colegio_id

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Omitir cabecera
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            raise HTTPException(status_code=400, detail="El archivo está vacío o no tiene datos.")
            
        def clean_time_str(val):
            if val is None:
                return None
            s = str(val).strip()
            return s[:5] if s else None

        data = []
        for row in rows[1:]:
            # Columnas esperadas: Nombre, ID Monitor, Cupos, Día, Hora Inicio, Hora Fin, Cursos Asignados, Año Académico, ID Colegio
            if not row or not row[0]:
                continue
                
            cursos_val = str(row[6]).strip() if len(row) > 6 and row[6] is not None else None
            periodo_val = None
            if len(row) > 7 and row[7] is not None:
                try:
                    periodo_val = int(row[7])
                except ValueError:
                    pass
            
            colegio_val = str(row[8]).strip() if len(row) > 8 and row[8] is not None else None
            if not colegio_val or colegio_val.lower() == "ninguno" or colegio_val == "":
                colegio_val = colegio_id

            data.append({
                "nombre_taller": str(row[0]).strip(),
                "profesor_id": str(row[1]).strip() if len(row) > 1 and row[1] is not None else "",
                "cupos_maximos": int(row[2]) if len(row) > 2 and row[2] is not None else 20,
                "dia": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "Lunes",
                "hora_inicio": clean_time_str(row[4]) if len(row) > 4 else None,
                "hora_fin": clean_time_str(row[5]) if len(row) > 5 else None,
                "cursos_asignados": cursos_val,
                "periodo": periodo_val,
                "colegio_id": colegio_val
            })
            
        stats = crud.bulk_upsert_talleres(db, data, colegio_id)
        return {"detail": "Importación completada", "stats": stats}
        
    except HTTPException:
        raise
    except Exception:
        logger.exception("Error procesando bulk-upload de talleres")
        raise HTTPException(status_code=500, detail="Error procesando el archivo")


@router.get("", response_model=List[TallerResponse])
def list_talleres(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    return crud.get_talleres(db, tenant.colegio_id, skip=skip, limit=limit, usuario_id=tenant.usuario_id, rol=tenant.rol)


@router.get("/otros-colegios", response_model=List[TallerResponse])
def list_talleres_otros_colegios(
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Sin permiso")
    return crud.get_talleres_otros_colegios(db, tenant.colegio_id)


@router.get("/{taller_id}", response_model=TallerResponse)
def get_taller(taller_id: UUID, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    db_taller = crud.get_taller_by_id(db, taller_id, tenant.colegio_id)
    if not db_taller:
        raise HTTPException(status_code=404, detail="Taller no encontrado")
    return db_taller


@router.post("", response_model=TallerResponse, status_code=status.HTTP_201_CREATED)
def create_taller(taller: TallerCreate, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para crear talleres")
    return crud.create_taller(db, taller, tenant.colegio_id)


@router.patch("/{taller_id}", response_model=TallerResponse)
def update_taller(taller_id: UUID, taller: TallerUpdate, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para editar talleres")
    db_taller = crud.update_taller(db, taller_id, taller, tenant.colegio_id)
    if not db_taller:
        raise HTTPException(status_code=404, detail="Taller no encontrado")
    return db_taller


@router.delete("/{taller_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_taller(taller_id: UUID, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para eliminar talleres")
    crud.delete_taller(db, taller_id, tenant.colegio_id)

@router.post("/clonar-periodo")
def clonar_periodo(data: TallerClonacion, db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="No tienes permisos para clonar periodos")
    
    clonados = crud.clonar_periodo(
        db, 
        data.source_periodo, 
        data.target_periodo, 
        data.taller_ids, 
        tenant.colegio_id
    )
    
    return {"detail": f"Se han clonado {clonados} talleres exitosamente para el periodo {data.target_periodo}"}

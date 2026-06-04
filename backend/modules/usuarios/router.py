from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from uuid import UUID
import io
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse
from core.database import get_db
from modules.auth.dependencies import get_current_tenant, TenantContext
from modules.usuarios.schemas import UsuarioCreate, UsuarioUpdate, UsuarioResponse, UsuarioBulkDelete
from modules.usuarios import crud

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/usuarios", tags=["usuarios"])

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.get("", response_model=List[UsuarioResponse])
def list_usuarios(
    skip: int = 0,
    limit: int = 100,
    role_filter: str = None,
    only_with_talleres: bool = False,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    return crud.get_usuarios(
        db,
        tenant.colegio_id,
        skip=skip,
        limit=limit,
        role_filter=role_filter,
        only_with_talleres=only_with_talleres,
        requester_role=tenant.rol,
        # El admin ve todos los usuarios (activos e inactivos)
        include_inactive=(tenant.rol == "admin")
    )


@router.post("/impacto-eliminacion")
def impacto_eliminacion(
    payload: UsuarioBulkDelete,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    if tenant.rol not in ("admin", "coordinador"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Sin permiso")
    return crud.get_impacto_eliminacion(db, payload.ids, tenant.colegio_id)


@router.post("/bulk-delete")
def bulk_delete_usuarios(
    payload: UsuarioBulkDelete,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    if tenant.rol not in ("admin", "coordinador"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo administradores o coordinadores pueden eliminar usuarios")
    if payload.permanente:
        if tenant.rol != "admin":
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo administradores pueden eliminar definitivamente")
        res = crud.hard_delete_usuarios_cascade(db, payload.ids, tenant.colegio_id, exclude_id=tenant.usuario_id)
        return {"detail": "Usuarios eliminados definitivamente", "permanente": True, "resultado": res}
    res = crud.bulk_delete_usuarios(db, payload.ids, tenant.colegio_id, exclude_id=tenant.usuario_id)
    return {"detail": "Usuarios eliminados", "permanente": False, "resultado": res}


@router.get("/template")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers_list = ["Nombre de Usuario (login)", "Nombre Completo", "Email", "Rol (admin/monitor/coordinador)", "Contraseña"]
    ws.append(headers_list)
    # Ejemplos
    ws.append(["jperez", "Juan Pérez González", "jperez@colegio.cl", "monitor", "jperez123"])
    ws.append(["mlopez", "María López Soto", "", "coordinador", ""])

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type=_XLSX_MEDIA, headers={
        'Content-Disposition': 'attachment; filename="planilla_usuarios.xlsx"'
    })


@router.get("/export")
def export_usuarios(db: Session = Depends(get_db), tenant: TenantContext = Depends(get_current_tenant)):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Sin permiso para exportar usuarios")

    usuarios = crud.get_usuarios(db, tenant.colegio_id, skip=0, limit=10000, requester_role=tenant.rol)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(["Nombre de Usuario", "Nombre Completo", "Email", "Rol", "Estado"])
    for u in usuarios:
        rol_val = u.rol.value if hasattr(u.rol, "value") else u.rol
        ws.append([u.nombre, u.nombre_2 or "", u.email or "", rol_val, "Activo" if u.is_active else "Inactivo"])

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type=_XLSX_MEDIA, headers={
        'Content-Disposition': 'attachment; filename="usuarios.xlsx"'
    })


@router.post("/bulk-upload")
def bulk_upload_usuarios(
    file: UploadFile = File(...),
    target_colegio_id: Optional[str] = None,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    if tenant.rol == "monitor":
        raise HTTPException(status_code=403, detail="Los monitores no tienen permiso para realizar cargas masivas")
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un .xlsx")

    colegio_id = tenant.colegio_id
    if target_colegio_id:
        if tenant.rol != "admin":
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo administradores pueden seleccionar un colegio de destino")
        colegio_id = target_colegio_id

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            raise HTTPException(status_code=400, detail="El archivo está vacío o no tiene datos.")

        data = []
        for row in rows[1:]:
            # Columnas: Login, Nombre Completo, Email, Rol, Contraseña
            if not row or not row[0]:
                continue
            data.append({
                "nombre": str(row[0]).strip(),
                "nombre_2": str(row[1]).strip() if len(row) > 1 and row[1] is not None else None,
                "email": str(row[2]).strip() if len(row) > 2 and row[2] is not None else None,
                "rol": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                "password": str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
            })

        stats = crud.bulk_create_usuarios(db, data, colegio_id)
        return {"detail": "Importación completada", "stats": stats}

    except HTTPException:
        raise
    except Exception:
        logger.exception("Error procesando bulk-upload de usuarios")
        raise HTTPException(status_code=500, detail="Error procesando el archivo")


@router.get("/{usuario_id}", response_model=UsuarioResponse)
def get_usuario(
    usuario_id: UUID,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    db_usuario = crud.get_usuario_by_id(db, usuario_id, tenant.colegio_id)
    if not db_usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return db_usuario


@router.post("", response_model=UsuarioResponse, status_code=status.HTTP_201_CREATED)
def create_usuario(
    usuario: UsuarioCreate,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    existing = crud.get_usuario_by_nombre(db, usuario.nombre, None)
    if existing:
        raise HTTPException(status_code=400, detail="El nombre de usuario ya está registrado")
    return crud.create_usuario(db, usuario, tenant.colegio_id)


@router.patch("/{usuario_id}", response_model=UsuarioResponse)
def update_usuario(
    usuario_id: UUID,
    usuario: UsuarioUpdate,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    try:
        db_usuario = crud.update_usuario(db, usuario_id, usuario, tenant.colegio_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not db_usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return db_usuario


@router.delete("/{usuario_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_usuario(
    usuario_id: UUID,
    db: Session = Depends(get_db),
    tenant: TenantContext = Depends(get_current_tenant)
):
    crud.delete_usuario(db, usuario_id, tenant.colegio_id)
